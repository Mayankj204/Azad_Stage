"""Azad Kishori — ALAP (Accelerator Leadership) sub-module routes.

ALAP is a child sub-module under the AK programme. It tracks individual
ALAP participants (a separate roster from AK Leaders), their community-
mapping activities, and their post-programme internship and employment
status.

Tables (mis_azad schema):
  - ak_alaps                   one row per ALAP participant
  - ak_alap_internships        many-per-alap, captures internship history
  - ak_alap_employment         many-per-alap (most recent typically wins)

Frontend lives under the AK programme's "ALAP" parent nav, expanded in
May-2026 to a 6-tab submenu (ALAP List + 5 placeholder tabs).
"""
from fastapi import APIRouter, HTTPException
from typing import Optional, List
from pydantic import BaseModel
from datetime import date
import sys, os, datetime
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from database import get_cursor

router = APIRouter(prefix="/api/ak/alap", tags=["AK ALAP"])


# ── Pydantic schemas ──────────────────────────────────────────────────────

class ALAPCreate(BaseModel):
    """Payload for create / update.

    Mirrors the FLP Save-as-Draft pattern: only `name` is strictly required
    on the model; everything else is Optional so a half-filled Draft can
    be saved. The frontend's full-Submit path enforces the wider set of
    mandatory fields described in the May-2026 spec (Name, Batch, DOB,
    Address, Education/Work Status, Family Members, Monthly Income).
    """
    name: str
    batch_id: Optional[int] = None
    state_code: Optional[str] = None
    centre_code: Optional[str] = None
    date_of_birth: Optional[date] = None
    address: Optional[str] = None
    category: Optional[str] = None
    category_other: Optional[str] = None
    community: Optional[str] = None
    community_other: Optional[str] = None
    education_work_status: Optional[str] = None
    education_work_other: Optional[str] = None
    family_members_count: Optional[int] = None
    monthly_family_income: Optional[float] = None
    mother_occupation: Optional[str] = None
    father_occupation: Optional[str] = None
    # Community Mapping section (tab 2 of the form)
    induction_start_date: Optional[date] = None
    induction_end_date: Optional[date] = None
    cm_location: Optional[str] = None
    cm_date: Optional[date] = None
    cm_infrastructure: Optional[str] = None
    status: Optional[str] = "Active"


class ALAPInternshipCreate(BaseModel):
    organization_name: Optional[str] = None
    sector: Optional[str] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None


class ALAPEmploymentCreate(BaseModel):
    is_employed: Optional[bool] = None
    employed_where: Optional[str] = None


# ── Enrollment-number generator ───────────────────────────────────────────

def _generate_enrollment(cur, batch_id: Optional[int]) -> str:
    """Build the canonical ALAP id of the form AF/ALAP/<Year>/<Serial>.

    Year is the financial year (Apr–Mar). Serial is one-based and unique
    within the year prefix. Batch is intentionally NOT in the prefix —
    serials run across the whole year so a batch swap doesn't reset the
    counter (matches the AK Leaders generator behaviour).

    2026-06-04 fix: scan BOTH live and soft-deleted rows when picking
    the max serial. The `ak_alaps.enrollment_number` UNIQUE constraint
    is column-wide and does NOT exempt soft-deleted rows (see
    `ak_alaps_enrollment_number_key` in pg_constraint). The previous
    code added `AND deleted_at IS NULL`, which caused this loop:
      1. Live max was 007; soft-deleted 008 existed in the table.
      2. Generator skipped 008 (filtered out), returned 008 as next.
      3. INSERT collided with the soft-deleted row, raising
         `duplicate key value violates unique constraint
         "ak_alaps_enrollment_number_key" DETAIL: Key
         (enrollment_number)=(AF/ALAP/2026-27/008) already exists.`
    Removing the `deleted_at` filter makes the generator agree with
    the constraint and the INSERT succeeds. Soft-deleted rows still
    burn a serial, which matches AK Leader behaviour and is the
    auditable choice (the enrollment number was issued; revoking it
    on delete would let it be reused under a different name, which
    is a worse outcome for traceability).
    """
    today = datetime.date.today()
    if today.month >= 4:
        year_str = f"{today.year}-{str(today.year+1)[-2:]}"
    else:
        year_str = f"{today.year-1}-{str(today.year)[-2:]}"
    prefix = f"AF/ALAP/{year_str}"
    cur.execute(
        "SELECT enrollment_number FROM ak_alaps "
        "WHERE enrollment_number LIKE %s "
        "ORDER BY enrollment_number DESC LIMIT 1",
        (prefix + "/%",),
    )
    row = cur.fetchone()
    if row and row.get("enrollment_number"):
        try:
            last = int(row["enrollment_number"].split("/")[-1])
            return f"{prefix}/{(last + 1):03d}"
        except (ValueError, IndexError):
            pass
    return f"{prefix}/001"


# ── List ──────────────────────────────────────────────────────────────────

@router.get("")
def list_alaps(
    batch_id: Optional[int] = None,
    name: Optional[str] = None,
    status: Optional[str] = None,
    state_code: Optional[str] = None,
    district_code: Optional[str] = None,
    centre_code: Optional[str] = None,
    page: int = 1, limit: int = 10,
):
    offset = max(0, (page - 1) * limit)
    conditions = ["a.deleted_at IS NULL"]
    params: list = []
    if batch_id:
        conditions.append("a.batch_id = %s"); params.append(batch_id)
    if name:
        conditions.append("LOWER(a.name) LIKE LOWER(%s)"); params.append(f"%{name}%")
    if status:
        conditions.append("a.status = %s"); params.append(status)
    if state_code:
        conditions.append("a.state_code = %s"); params.append(state_code)
    if district_code:
        # ak_alaps carries only state/centre — expand to all centres in district.
        conditions.append("a.centre_code IN (SELECT centre_code FROM ak_centres WHERE district_code = %s)")
        params.append(district_code)
    if centre_code:
        conditions.append("a.centre_code = %s"); params.append(centre_code)

    where_sql = " AND ".join(conditions)
    with get_cursor() as cur:
        cur.execute(f"SELECT COUNT(*) AS total FROM ak_alaps a WHERE {where_sql}", params)
        total = cur.fetchone()["total"]
        cur.execute(
            f"""
            SELECT a.id, a.enrollment_number, a.name, a.batch_id, b.name AS batch_name,
                   a.state_code, a.centre_code,
                   COALESCE(ns.state_name, '')  AS state_name,
                   COALESCE(nc.centre_name, '') AS centre_name,
                   a.education_work_status, a.induction_start_date, a.induction_end_date,
                   a.status, a.created_at
            FROM ak_alaps a
            LEFT JOIN ak_batches b   ON a.batch_id    = b.id
            LEFT JOIN ak_states  ns  ON a.state_code  = ns.state_code
            LEFT JOIN ak_centres nc  ON a.centre_code = nc.centre_code
            WHERE {where_sql}
            ORDER BY a.created_at DESC, a.id DESC
            LIMIT %s OFFSET %s
            """,
            params + [limit, offset],
        )
        rows = cur.fetchall()
    return {"total": total, "page": page, "limit": limit, "data": rows}


# ── Create ────────────────────────────────────────────────────────────────

@router.post("")
def create_alap(alap: ALAPCreate):
    with get_cursor() as cur:
        enrollment = _generate_enrollment(cur, alap.batch_id)
        try:
            cur.execute(
                """
                INSERT INTO ak_alaps (
                    enrollment_number, name, batch_id, state_code, centre_code,
                    date_of_birth, address, category, category_other,
                    community, community_other, education_work_status,
                    education_work_other, family_members_count,
                    monthly_family_income, mother_occupation, father_occupation,
                    induction_start_date, induction_end_date,
                    cm_location, cm_date, cm_infrastructure, status
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id, enrollment_number, name
                """,
                (
                    enrollment, alap.name, alap.batch_id, alap.state_code, alap.centre_code,
                    alap.date_of_birth, alap.address, alap.category, alap.category_other,
                    alap.community, alap.community_other, alap.education_work_status,
                    alap.education_work_other, alap.family_members_count,
                    alap.monthly_family_income, alap.mother_occupation, alap.father_occupation,
                    alap.induction_start_date, alap.induction_end_date,
                    alap.cm_location, alap.cm_date, alap.cm_infrastructure, alap.status or "Active",
                ),
            )
            new = cur.fetchone()
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

        try:
            cur.execute(
                """
                INSERT INTO system_activity_log (user_name, action, resource_type, resource_id, description, source)
                VALUES (%s, 'Create ALAP', 'ALAP', %s, %s, 'web')
                """,
                (alap.name, new["id"], f'ALAP {alap.name} ({new["enrollment_number"]}) created'),
            )
        except Exception:
            pass
        return new


# ── Detail ────────────────────────────────────────────────────────────────

@router.get("/{alap_id}")
def get_alap(alap_id: int):
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT a.*, b.name AS batch_name,
                   COALESCE(ns.state_name, '') AS state_name,
                   COALESCE(nc.centre_name, '') AS centre_name
            FROM ak_alaps a
            LEFT JOIN ak_batches b   ON a.batch_id = b.id
            LEFT JOIN new_states ns  ON a.state_code = ns.state_code
            LEFT JOIN new_centres nc ON a.centre_code = nc.centre_code
            WHERE a.id = %s AND a.deleted_at IS NULL
            """,
            (alap_id,),
        )
        alap = cur.fetchone()
        if not alap:
            raise HTTPException(status_code=404, detail="ALAP not found")

        # Pull internship + employment history.
        cur.execute(
            "SELECT * FROM ak_alap_internships WHERE alap_id = %s ORDER BY start_date DESC NULLS LAST, id DESC",
            (alap_id,),
        )
        internships = cur.fetchall()
        cur.execute(
            "SELECT * FROM ak_alap_employment WHERE alap_id = %s ORDER BY id DESC",
            (alap_id,),
        )
        employment = cur.fetchall()

    out = dict(alap)
    out["internships"] = internships
    out["employment"] = employment
    return out


# ── Update ────────────────────────────────────────────────────────────────

@router.put("/{alap_id}")
def update_alap(alap_id: int, alap: ALAPCreate):
    with get_cursor() as cur:
        cur.execute(
            """
            UPDATE ak_alaps SET
                name=%s, batch_id=%s, state_code=%s, centre_code=%s,
                date_of_birth=%s, address=%s, category=%s, category_other=%s,
                community=%s, community_other=%s, education_work_status=%s,
                education_work_other=%s, family_members_count=%s,
                monthly_family_income=%s, mother_occupation=%s, father_occupation=%s,
                induction_start_date=%s, induction_end_date=%s,
                cm_location=%s, cm_date=%s, cm_infrastructure=%s,
                status=%s, updated_at=NOW()
            WHERE id=%s AND deleted_at IS NULL RETURNING id
            """,
            (
                alap.name, alap.batch_id, alap.state_code, alap.centre_code,
                alap.date_of_birth, alap.address, alap.category, alap.category_other,
                alap.community, alap.community_other, alap.education_work_status,
                alap.education_work_other, alap.family_members_count,
                alap.monthly_family_income, alap.mother_occupation, alap.father_occupation,
                alap.induction_start_date, alap.induction_end_date,
                alap.cm_location, alap.cm_date, alap.cm_infrastructure,
                alap.status or "Active", alap_id,
            ),
        )
        result = cur.fetchone()
        if not result:
            raise HTTPException(status_code=404, detail="ALAP not found")
        try:
            cur.execute(
                """
                INSERT INTO system_activity_log (user_name, action, resource_type, resource_id, description, source)
                VALUES (%s, 'Update ALAP', 'ALAP', %s, %s, 'web')
                """,
                (alap.name, alap_id, f"ALAP {alap.name} (ID:{alap_id}) updated"),
            )
        except Exception:
            pass
        return {"success": True, "id": alap_id}


# ── Soft delete ───────────────────────────────────────────────────────────

@router.delete("/{alap_id}")
def delete_alap(alap_id: int):
    with get_cursor() as cur:
        cur.execute("UPDATE ak_alaps SET deleted_at = NOW() WHERE id = %s RETURNING id", (alap_id,))
        result = cur.fetchone()
        if not result:
            raise HTTPException(status_code=404, detail="ALAP not found")
        return {"success": True, "id": alap_id}


# ── Internship sub-resource ───────────────────────────────────────────────

@router.post("/{alap_id}/internship")
def add_internship(alap_id: int, body: ALAPInternshipCreate):
    with get_cursor() as cur:
        # Sanity: confirm the parent ALAP exists.
        cur.execute("SELECT id FROM ak_alaps WHERE id=%s AND deleted_at IS NULL", (alap_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="ALAP not found")
        cur.execute(
            """
            INSERT INTO ak_alap_internships (alap_id, organization_name, sector, start_date, end_date)
            VALUES (%s,%s,%s,%s,%s) RETURNING id
            """,
            (alap_id, body.organization_name, body.sector, body.start_date, body.end_date),
        )
        return cur.fetchone()


@router.get("/{alap_id}/internship")
def list_internships(alap_id: int):
    with get_cursor() as cur:
        cur.execute(
            "SELECT * FROM ak_alap_internships WHERE alap_id=%s ORDER BY start_date DESC NULLS LAST, id DESC",
            (alap_id,),
        )
        return cur.fetchall()


# ── Employment sub-resource ───────────────────────────────────────────────

@router.post("/{alap_id}/employment")
def add_employment(alap_id: int, body: ALAPEmploymentCreate):
    with get_cursor() as cur:
        cur.execute("SELECT id FROM ak_alaps WHERE id=%s AND deleted_at IS NULL", (alap_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="ALAP not found")
        cur.execute(
            """
            INSERT INTO ak_alap_employment (alap_id, is_employed, employed_where)
            VALUES (%s,%s,%s) RETURNING id
            """,
            (alap_id, body.is_employed, body.employed_where),
        )
        return cur.fetchone()


@router.get("/{alap_id}/employment")
def list_employment(alap_id: int):
    with get_cursor() as cur:
        cur.execute(
            "SELECT * FROM ak_alap_employment WHERE alap_id=%s ORDER BY id DESC",
            (alap_id,),
        )
        return cur.fetchall()


# ── Excel export ──────────────────────────────────────────────────────────

@router.get("/export/excel")
def export_alaps_excel(
    batch_id: Optional[int] = None, name: Optional[str] = None,
    status: Optional[str] = None, state_code: Optional[str] = None,
    district_code: Optional[str] = None, centre_code: Optional[str] = None,
):
    """Stream the filtered ALAP list as an .xlsx file via openpyxl.

    Mirrors the FLP / AK list exports — same header style, same colour,
    same date formatting. Filters cascade from the same params as
    GET /api/ak/alap so the user gets exactly what they see in the table.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    import io as _io

    conditions = ["a.deleted_at IS NULL"]
    params: list = []
    if batch_id:
        conditions.append("a.batch_id = %s"); params.append(batch_id)
    if name:
        conditions.append("LOWER(a.name) LIKE LOWER(%s)"); params.append(f"%{name}%")
    if status:
        conditions.append("a.status = %s"); params.append(status)
    if state_code:
        conditions.append("a.state_code = %s"); params.append(state_code)
    if district_code:
        conditions.append("a.centre_code IN (SELECT centre_code FROM ak_centres WHERE district_code = %s)")
        params.append(district_code)
    if centre_code:
        conditions.append("a.centre_code = %s"); params.append(centre_code)
    where_sql = " AND ".join(conditions)

    with get_cursor() as cur:
        cur.execute(
            f"""
            SELECT a.id, a.enrollment_number, a.name, b.name AS batch_name,
                   a.date_of_birth, a.education_work_status,
                   a.induction_start_date, a.induction_end_date,
                   a.cm_location, a.status, a.created_at
            FROM ak_alaps a LEFT JOIN ak_batches b ON a.batch_id = b.id
            WHERE {where_sql}
            ORDER BY a.created_at DESC, a.id DESC
            """,
            params,
        )
        rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "ALAP List"
    headers = ["S.No", "Enrollment No.", "Name", "Batch", "Date of Birth",
               "Education / Work Status", "Induction Start", "Induction End",
               "Location", "Status", "Created At"]
    ws.append(headers)
    header_fill = PatternFill(start_color="732269", end_color="732269", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, start=1):
        ws.append([
            i,
            r.get("enrollment_number") or "",
            r.get("name") or "",
            r.get("batch_name") or "",
            r["date_of_birth"].isoformat() if r.get("date_of_birth") else "",
            r.get("education_work_status") or "",
            r["induction_start_date"].isoformat() if r.get("induction_start_date") else "",
            r["induction_end_date"].isoformat() if r.get("induction_end_date") else "",
            r.get("cm_location") or "",
            r.get("status") or "",
            r["created_at"].strftime("%Y-%m-%d %H:%M") if r.get("created_at") else "",
        ])

    # Auto-fit column widths (rough heuristic).
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(14, len(h) + 4)

    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"ALAP_List_{datetime.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
