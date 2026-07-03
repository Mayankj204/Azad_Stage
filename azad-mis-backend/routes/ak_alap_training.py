"""ALAP Training sub-module routes.

Sits inside the AK programme alongside ak_alap. Manages training events
that ALAP participants attend — each event has a state, a phase (I-IV),
a date range, and a multi-select set of training-types (Residential
Training, Monthly Input Session, Exposure Visit, Monthly Meeting,
Community Action Project, Campaign), plus type-conditional detail fields
(topic / action point / campaign name / etc.).

Three tables on the mis_azad schema:
  - ak_alap_trainings              one row per training event
  - ak_alap_training_assignments   many-to-many: which ALAPs got assigned
  - ak_alap_training_attendance    per (training, alap, sub-type, month)
                                   — one training can have many attendance
                                   sessions across months / sub-types.

Attendance status flow: Draft → Submitted (saved by the user). Drafts are
overwritten on subsequent saves; Submitted records can be re-saved
without state-machine constraints in this initial cut.
"""
from fastapi import APIRouter, HTTPException
from typing import Optional, List
from pydantic import BaseModel
from datetime import date
import sys, os, json
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from database import get_cursor

router = APIRouter(prefix="/api/ak/alap-training", tags=["AK ALAP Training"])


class TrainingCreate(BaseModel):
    state_code: Optional[str] = None
    phase: Optional[str] = None
    # Start/End date columns are kept on the schema for legacy data but
    # no longer captured on the form (May-2026 spec change). The form
    # now uses `month` as the single time anchor.
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    month: Optional[str] = None
    # Multi-select set of training-type labels picked on the form. Stored
    # as a JSON array so we don't need a parallel mapping table for what
    # is essentially a checkbox group.
    training_types: Optional[List[str]] = None
    # Per-type detail bundle. Keyed by training type name (e.g.
    # "Residential Training") with each value being an object of that
    # type's fields. This avoids the previous "shared topic/duration
    # columns get overwritten by whichever type is filled last" bug.
    # Stored as JSONB on ak_alap_trainings.type_details.
    type_details: Optional[dict] = None
    # Legacy flat columns — kept Optional for back-compat. New writes
    # only populate type_details; the flat columns stay null. Reads
    # fall back to these when type_details is missing (older rows).
    topic: Optional[str] = None
    detail_date: Optional[date] = None
    action_point: Optional[str] = None
    duration: Optional[str] = None
    cap_details: Optional[str] = None
    campaign_name: Optional[str] = None
    campaign_details: Optional[str] = None
    status: Optional[str] = "Active"


class AssignmentBatch(BaseModel):
    """Body for the Assign-members modal save.

    `alap_ids` is the FULL desired set of assigned ALAPs after the user
    closes the modal. The route diff-replaces — adds new IDs, removes
    those that were unchecked. Simpler from the UI's perspective than
    sending diffs.
    """
    alap_ids: List[int]


class AttendanceEntry(BaseModel):
    alap_id: int
    present: bool


class AttendanceBatch(BaseModel):
    """Body for the Take-Attendance modal save.

    `training_subtype` and `month` together define the attendance scope —
    one training can have many attendance sessions over its lifecycle
    (e.g. attendance for "Residential Training" in May, then again for
    "Monthly Meeting" in June).

    `status` flips between 'Draft' (Save Draft button) and 'Submitted'
    (Submit Attendance button). Re-saves overwrite.
    """
    training_subtype: Optional[str] = None
    attendance_month: Optional[str] = None
    status: str = "Submitted"
    entries: List[AttendanceEntry]


# ── List / pagination ─────────────────────────────────────────────────────

@router.get("")
def list_trainings(
    state_code: Optional[str] = None,
    phase: Optional[str] = None,
    month: Optional[str] = None,
    # Legacy date_from/date_to params still accepted for backward-compat
    # callers but the frontend stopped sending them after the May-2026
    # spec change that replaced the duration date-range with a Month
    # filter.
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    page: int = 1, limit: int = 10,
):
    offset = max(0, (page - 1) * limit)
    conditions = ["t.deleted_at IS NULL"]
    params: list = []
    if state_code:
        conditions.append("t.state_code = %s"); params.append(state_code)
    if phase:
        conditions.append("t.phase = %s"); params.append(phase)
    if month:
        conditions.append("t.month = %s"); params.append(month)
    if date_from:
        conditions.append("t.start_date >= %s"); params.append(date_from)
    if date_to:
        conditions.append("t.end_date <= %s"); params.append(date_to)
    where_sql = " AND ".join(conditions)

    with get_cursor() as cur:
        cur.execute(f"SELECT COUNT(*) AS total FROM ak_alap_trainings t WHERE {where_sql}", params)
        total = cur.fetchone()["total"]
        cur.execute(
            f"""
            SELECT t.id, t.state_code, t.phase, t.start_date, t.end_date, t.month,
                   t.training_types, t.topic, t.detail_date, t.action_point,
                   t.duration, t.cap_details, t.campaign_name, t.campaign_details,
                   t.status, t.created_at,
                   COALESCE(ns.state_name, '') AS state_name,
                   (SELECT COUNT(*) FROM ak_alap_training_assignments a WHERE a.training_id = t.id) AS participants
            FROM ak_alap_trainings t
            LEFT JOIN ak_states  ns ON t.state_code = ns.state_code
            WHERE {where_sql}
            ORDER BY t.start_date DESC NULLS LAST, t.id DESC
            LIMIT %s OFFSET %s
            """,
            params + [limit, offset],
        )
        rows = cur.fetchall()
    return {"total": total, "page": page, "limit": limit, "data": rows}


# ── Create ─────────────────────────────────────────────────────────────────

@router.post("")
def create_training(t: TrainingCreate):
    with get_cursor() as cur:
        cur.execute(
            """
            INSERT INTO ak_alap_trainings (
                state_code, phase, start_date, end_date, month,
                training_types, type_details, topic, detail_date, action_point,
                duration, cap_details, campaign_name, campaign_details, status
            ) VALUES (%s,%s,%s,%s,%s,%s::jsonb,%s::jsonb,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id, state_code, phase, month
            """,
            (
                t.state_code, t.phase, t.start_date, t.end_date, t.month,
                json.dumps(t.training_types or []),
                json.dumps(t.type_details or {}),
                t.topic, t.detail_date, t.action_point,
                t.duration, t.cap_details, t.campaign_name, t.campaign_details,
                t.status or "Active",
            ),
        )
        return cur.fetchone()


# ── Detail ─────────────────────────────────────────────────────────────────

@router.get("/{training_id}")
def get_training(training_id: int):
    with get_cursor() as cur:
        cur.execute(
            """
            SELECT t.*, COALESCE(ns.state_name, '') AS state_name
            FROM ak_alap_trainings t
            LEFT JOIN ak_states  ns ON t.state_code = ns.state_code
            WHERE t.id = %s AND t.deleted_at IS NULL
            """,
            (training_id,),
        )
        t = cur.fetchone()
        if not t:
            raise HTTPException(status_code=404, detail="Training not found")

        cur.execute(
            """
            SELECT a.alap_id, ap.name AS alap_name, ap.enrollment_number,
                   b.name AS batch_name, ap.status AS alap_status, a.assigned_at
            FROM ak_alap_training_assignments a
            JOIN ak_alaps  ap ON a.alap_id = ap.id
            LEFT JOIN ak_batches b ON ap.batch_id = b.id
            WHERE a.training_id = %s AND ap.deleted_at IS NULL
            ORDER BY ap.name
            """,
            (training_id,),
        )
        assigned = cur.fetchall()

        # Attendance counts grouped by training_subtype so the View page
        # can render "X / Y present" beside each type's detail block.
        # Total = number of assigned ALAPs at the time of query (not the
        # number who have records for this subtype) — matches the
        # denominator the user expects on a per-type breakdown.
        cur.execute(
            """
            SELECT training_subtype,
                   COUNT(*) FILTER (WHERE present = TRUE)  AS present_count,
                   COUNT(*)                                AS recorded_count
            FROM ak_alap_training_attendance
            WHERE training_id = %s AND training_subtype IS NOT NULL
            GROUP BY training_subtype
            """,
            (training_id,),
        )
        rows = cur.fetchall()
        attendance_counts = {}
        for r in rows:
            attendance_counts[r["training_subtype"]] = {
                "present": r["present_count"] or 0,
                "recorded": r["recorded_count"] or 0,
            }

        # Per-subtype attendee breakdown — used by the View page to list
        # who actually attended each training type (not just the count).
        cur.execute(
            """
            SELECT att.training_subtype,
                   ap.id   AS alap_id,
                   ap.name AS alap_name,
                   ap.enrollment_number,
                   b.name  AS batch_name,
                   att.present,
                   att.recorded_at
            FROM ak_alap_training_attendance att
            JOIN ak_alaps ap ON att.alap_id = ap.id
            LEFT JOIN ak_batches b ON ap.batch_id = b.id
            WHERE att.training_id = %s AND att.training_subtype IS NOT NULL
            ORDER BY att.training_subtype, ap.name
            """,
            (training_id,),
        )
        attendance_breakdown: dict = {}
        for r in cur.fetchall():
            attendance_breakdown.setdefault(r["training_subtype"], []).append({
                "alap_id": r["alap_id"],
                "alap_name": r["alap_name"],
                "enrollment_number": r["enrollment_number"],
                "batch_name": r["batch_name"],
                "present": r["present"],
            })
    out = dict(t)
    out["assigned"] = assigned
    out["attendance_counts"] = attendance_counts
    out["attendance_breakdown"] = attendance_breakdown
    out["assigned_total"] = len(assigned)
    return out


# ── Update ─────────────────────────────────────────────────────────────────

@router.put("/{training_id}")
def update_training(training_id: int, t: TrainingCreate):
    with get_cursor() as cur:
        cur.execute(
            """
            UPDATE ak_alap_trainings SET
                state_code=%s, phase=%s, start_date=%s, end_date=%s, month=%s,
                training_types=%s::jsonb, type_details=%s::jsonb,
                topic=%s, detail_date=%s, action_point=%s,
                duration=%s, cap_details=%s, campaign_name=%s, campaign_details=%s,
                status=%s, updated_at=NOW()
            WHERE id=%s AND deleted_at IS NULL RETURNING id
            """,
            (
                t.state_code, t.phase, t.start_date, t.end_date, t.month,
                json.dumps(t.training_types or []),
                json.dumps(t.type_details or {}),
                t.topic, t.detail_date, t.action_point,
                t.duration, t.cap_details, t.campaign_name, t.campaign_details,
                t.status or "Active", training_id,
            ),
        )
        result = cur.fetchone()
        if not result:
            raise HTTPException(status_code=404, detail="Training not found")
        return {"success": True, "id": training_id}


# ── Soft delete ────────────────────────────────────────────────────────────

@router.delete("/{training_id}")
def delete_training(training_id: int):
    with get_cursor() as cur:
        cur.execute("UPDATE ak_alap_trainings SET deleted_at=NOW() WHERE id=%s RETURNING id", (training_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Training not found")
        return {"success": True, "id": training_id}


# ── Assignments (Assign Members modal) ─────────────────────────────────────

@router.get("/{training_id}/assignments")
def list_assignments(training_id: int):
    """Return the FULL ALAP roster + a flag indicating who's currently
    assigned to this training. The Assign modal renders this list with
    pre-checked rows for the assigned ones — saves a separate fetch.
    """
    with get_cursor() as cur:
        # Confirm training exists.
        cur.execute("SELECT id FROM ak_alap_trainings WHERE id=%s AND deleted_at IS NULL", (training_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Training not found")
        cur.execute(
            """
            SELECT ap.id, ap.name, ap.enrollment_number, ap.status,
                   b.name AS batch_name,
                   EXISTS (
                       SELECT 1 FROM ak_alap_training_assignments a
                       WHERE a.training_id = %s AND a.alap_id = ap.id
                   ) AS assigned
            FROM ak_alaps ap
            LEFT JOIN ak_batches b ON ap.batch_id = b.id
            WHERE ap.deleted_at IS NULL AND ap.status = 'Active'
            ORDER BY ap.name
            """,
            (training_id,),
        )
        return cur.fetchall()


@router.put("/{training_id}/assignments")
def set_assignments(training_id: int, body: AssignmentBatch):
    """Replace the assignments for this training with the given set.

    Diff-replaces — adds rows for newly-checked alap_ids, deletes rows
    for the unchecked ones. Inside a transaction so a partial failure
    doesn't leave the assignment list in a half-committed state.
    """
    with get_cursor() as cur:
        cur.execute("SELECT id FROM ak_alap_trainings WHERE id=%s AND deleted_at IS NULL", (training_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Training not found")
        cur.execute("SELECT alap_id FROM ak_alap_training_assignments WHERE training_id=%s", (training_id,))
        existing = {r["alap_id"] for r in cur.fetchall()}
        wanted  = set(body.alap_ids or [])
        to_add  = wanted - existing
        to_drop = existing - wanted
        for aid in to_add:
            cur.execute(
                "INSERT INTO ak_alap_training_assignments (training_id, alap_id) VALUES (%s,%s) "
                "ON CONFLICT (training_id, alap_id) DO NOTHING",
                (training_id, aid),
            )
        if to_drop:
            cur.execute(
                "DELETE FROM ak_alap_training_assignments WHERE training_id=%s AND alap_id = ANY(%s)",
                (training_id, list(to_drop)),
            )
        return {"success": True, "added": len(to_add), "removed": len(to_drop)}


# ── Attendance (Take Attendance modal) ─────────────────────────────────────

@router.get("/{training_id}/attendance")
def get_attendance(
    training_id: int,
    training_subtype: Optional[str] = None,
    attendance_month: Optional[str] = None,
):
    """Return the assigned roster + their attendance for a (subtype, month)
    pair. If the user hasn't recorded attendance for this combo yet,
    every row comes back with present=null so the modal can render empty
    checkboxes.
    """
    with get_cursor() as cur:
        cur.execute("SELECT id FROM ak_alap_trainings WHERE id=%s AND deleted_at IS NULL", (training_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Training not found")
        # Pull the assigned roster + any existing attendance row for the
        # requested subtype/month combo via a LEFT JOIN. The COALESCE in
        # the SELECT keeps the response shape stable when no attendance
        # has been recorded yet.
        cur.execute(
            """
            SELECT ap.id AS alap_id, ap.name, ap.enrollment_number, ap.status,
                   b.name AS batch_name,
                   att.present, att.attendance_status, att.recorded_at
            FROM ak_alap_training_assignments a
            JOIN ak_alaps ap ON a.alap_id = ap.id AND ap.deleted_at IS NULL
            LEFT JOIN ak_batches b ON ap.batch_id = b.id
            LEFT JOIN ak_alap_training_attendance att
              ON att.training_id = a.training_id
             AND att.alap_id     = a.alap_id
             AND att.training_subtype IS NOT DISTINCT FROM %s
             AND att.attendance_month IS NOT DISTINCT FROM %s
            WHERE a.training_id = %s
            ORDER BY ap.name
            """,
            (training_subtype, attendance_month, training_id),
        )
        return cur.fetchall()


@router.post("/{training_id}/attendance")
def save_attendance(training_id: int, body: AttendanceBatch):
    """Upsert the attendance batch.

    Each entry creates or updates a row keyed on (training, alap, subtype,
    month). The status field flips between 'Draft' and 'Submitted' so the
    Save Draft / Submit Attendance buttons share this same endpoint.
    """
    if body.status not in ("Draft", "Submitted"):
        raise HTTPException(status_code=400, detail="status must be 'Draft' or 'Submitted'")
    with get_cursor() as cur:
        cur.execute("SELECT id FROM ak_alap_trainings WHERE id=%s AND deleted_at IS NULL", (training_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Training not found")
        for e in body.entries:
            cur.execute(
                """
                INSERT INTO ak_alap_training_attendance
                  (training_id, alap_id, training_subtype, attendance_month,
                   present, attendance_status, recorded_at)
                VALUES (%s,%s,%s,%s,%s,%s,NOW())
                ON CONFLICT (training_id, alap_id, training_subtype, attendance_month)
                DO UPDATE SET
                  present           = EXCLUDED.present,
                  attendance_status = EXCLUDED.attendance_status,
                  recorded_at       = NOW()
                """,
                (training_id, e.alap_id, body.training_subtype, body.attendance_month,
                 bool(e.present), body.status),
            )
        return {"success": True, "saved": len(body.entries), "status": body.status}


# ── Excel export ───────────────────────────────────────────────────────────

@router.get("/export/excel")
def export_trainings_excel(
    state_code: Optional[str] = None,
    phase: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    import io as _io, datetime as _dt

    conditions = ["t.deleted_at IS NULL"]
    params: list = []
    if state_code: conditions.append("t.state_code = %s"); params.append(state_code)
    if phase:      conditions.append("t.phase = %s");      params.append(phase)
    if date_from:  conditions.append("t.start_date >= %s"); params.append(date_from)
    if date_to:    conditions.append("t.end_date <= %s");   params.append(date_to)
    where_sql = " AND ".join(conditions)

    with get_cursor() as cur:
        cur.execute(
            f"""
            SELECT t.id, COALESCE(ns.state_name, t.state_code) AS state_name,
                   t.phase, t.start_date, t.end_date, t.month,
                   t.training_types, t.topic, t.detail_date,
                   t.action_point, t.duration, t.cap_details,
                   t.campaign_name, t.campaign_details, t.status,
                   (SELECT COUNT(*) FROM ak_alap_training_assignments a WHERE a.training_id = t.id) AS participants
            FROM ak_alap_trainings t
            LEFT JOIN ak_states  ns ON t.state_code = ns.state_code
            WHERE {where_sql}
            ORDER BY t.start_date DESC NULLS LAST, t.id DESC
            """,
            params,
        )
        rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "ALAP Training"
    headers = ["S.No", "State", "Phase", "Start Date", "End Date", "Month",
               "Training Types", "Topic", "Action Point", "Campaign Name",
               "Status", "Participants"]
    ws.append(headers)
    fill = PatternFill(start_color="732269", end_color="732269", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, start=1):
        types = r.get("training_types") or []
        if isinstance(types, str):
            try: types = json.loads(types)
            except Exception: types = []
        ws.append([
            i,
            r.get("state_name") or "",
            r.get("phase") or "",
            r["start_date"].isoformat() if r.get("start_date") else "",
            r["end_date"].isoformat() if r.get("end_date") else "",
            r.get("month") or "",
            ", ".join(types) if types else "",
            r.get("topic") or "",
            r.get("action_point") or "",
            r.get("campaign_name") or "",
            r.get("status") or "",
            r.get("participants") or 0,
        ])
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(14, len(h) + 4)

    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"ALAP_Training_{_dt.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
