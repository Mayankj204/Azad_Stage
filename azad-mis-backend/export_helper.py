"""Utility to convert CSV-style data to XLSX response.

Centralizing export logic here ensures every downloadable report is a real
xlsx binary (fixes "file format or file extension is not valid" errors that
appeared when CSV content was served with .xlsx extension + xlsx MIME type).
"""
import io
import csv
from openpyxl import Workbook
from fastapi.responses import StreamingResponse
from datetime import date as _date_t, datetime as _datetime_t

# 2026-06-25: client-requested behaviour — every date cell in every export
# should render as a real Excel date (so the auto-filter groups by Year >
# Month > Day) and display in dd-mm-yy form, matching the Indian convention.
_DATE_FORMAT = 'dd-mm-yy'

def _coerce_date_value(v):
    """If v looks like a calendar date, return a datetime.date.  Else None.

    Conservative: only matches strict YYYY-MM-DD and DD-MM-YYYY shapes, and
    datetime/date objects.  Random strings that happen to contain digits and
    dashes (e.g. 'S07091B2') stay as text.
    """
    if v is None or v == '':
        return None
    if isinstance(v, _datetime_t):
        return v.date()
    if isinstance(v, _date_t):
        return v
    if not isinstance(v, str):
        return None
    s = v.strip()
    if not s:
        return None
    # Strip time portion if present.
    if 'T' in s:
        s = s.split('T', 1)[0]
    elif ' ' in s and len(s) >= 10:
        # 'YYYY-MM-DD HH:MM:SS...' or 'DD-MM-YYYY HH:MM...'
        s = s[:10]
    if len(s) != 10:
        return None
    try:
        # YYYY-MM-DD
        if s[4] == '-' and s[7] == '-':
            return _date_t(int(s[0:4]), int(s[5:7]), int(s[8:10]))
        # DD-MM-YYYY
        if s[2] == '-' and s[5] == '-':
            d, m, y = s.split('-')
            return _date_t(int(y), int(m), int(d))
    except (ValueError, TypeError):
        return None
    return None


def _write_cell(ws, row, col, v):
    """Write v into ws[row,col], converting date-shaped values to true date cells."""
    cell = ws.cell(row=row, column=col)
    d = _coerce_date_value(v)
    if d is not None:
        cell.value = d
        cell.number_format = _DATE_FORMAT
    else:
        cell.value = '' if v is None else v




def _build_xlsx_response(headers, rows, filename):
    """Shared: write headers + rows into a real xlsx workbook and return a StreamingResponse."""
    wb = Workbook()
    ws = wb.active
    if headers:
        ws.append(list(headers))
    # Use _write_cell so date-shaped values become real Excel dates (dd-mm-yy)
    header_offset = 1 if headers else 0
    for i, row in enumerate(rows):
        for j, v in enumerate(row, start=1):
            _write_cell(ws, header_offset + 1 + i, j, v)
    # Auto-width columns (bounded so very long text doesn't create absurd widths)
    for col in ws.columns:
        try:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        except Exception:
            pass
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    # Ensure filename ends with .xlsx (caller may pass .csv by mistake from legacy code)
    if not filename.lower().endswith('.xlsx'):
        filename = filename.rsplit('.', 1)[0] + '.xlsx'
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


def csv_to_xlsx_response(headers, rows, filename):
    """Convert pre-built headers + rows to an xlsx StreamingResponse."""
    return _build_xlsx_response(headers, rows, filename)


def csv_string_to_xlsx_response(csv_text, filename):
    """Parse an in-memory CSV string and return a real xlsx StreamingResponse.

    Drop-in replacement for endpoints that already wrote CSV via csv.writer into
    a StringIO — they can now do:
        return csv_string_to_xlsx_response(output.getvalue(), 'Export.xlsx')
    """
    if csv_text and csv_text.startswith('\ufeff'):
        csv_text = csv_text[1:]  # strip BOM if present
    reader = csv.reader(io.StringIO(csv_text))
    rows_iter = list(reader)
    headers = rows_iter[0] if rows_iter else []
    data_rows = rows_iter[1:] if len(rows_iter) > 1 else []
    return _build_xlsx_response(headers, data_rows, filename)


def multi_sheet_xlsx_response(sheets, filename):
    """Build a single .xlsx workbook with multiple sheets.

    sheets: list of tuples (sheet_name: str, headers: list, rows: iterable of lists)
    """
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name, headers, rows in sheets:
        safe = (sheet_name or 'Sheet')[:31]
        for bad in ('\\', '/', '*', '?', '[', ']', ':'):
            safe = safe.replace(bad, '-')
        ws = wb.create_sheet(title=safe)
        if headers:
            ws.append(list(headers))
        header_offset = 1 if headers else 0
        for i, row in enumerate(rows):
            for j, v in enumerate(row, start=1):
                _write_cell(ws, header_offset + 1 + i, j, v)
        for col in ws.columns:
            try:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
            except Exception:
                pass

    if not wb.sheetnames:
        wb.create_sheet(title='Empty')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    if not filename.lower().endswith('.xlsx'):
        filename = filename.rsplit('.', 1)[0] + '.xlsx'
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


def multi_sheet_xlsx_response_v2(sheets, filename):
    """Build a single .xlsx with multi-row headers and optional cell merges.

    Each `sheets` entry is a dict:
        {
          'name': <sheet name>,
          'group_headers': [(start_col, end_col, title), ...]  # optional row-1 groups
          'headers': [...flat column names...]                  # row-2 (or row-1 if no groups)
          'rows': iterable of lists
        }
    Column indices are 1-based for `group_headers`.
    """
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    group_font = Font(bold=True, size=11, color='FFFFFFFF')
    # Palette — each merged group cycles through a distinct color so sections
    # are visually separated. Mirrors the reference FLP MIS workbook style.
    group_palette = [
        '732269',  # purple
        '3498DB',  # blue
        '27AE60',  # green
        'E67E22',  # orange
        'C0392B',  # red
        '8E44AD',  # violet
        '16A085',  # teal
        'D35400',  # dark orange
        '2C3E50',  # navy
        '7F8C8D',  # slate
        'B03A2E',  # brick
        '1ABC9C',  # mint
    ]
    header_font = Font(bold=True, size=10, color='FF333333')
    header_fill = PatternFill(start_color='F3E6F0', end_color='F3E6F0', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='FFCCCCCC')
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet in sheets:
        name = sheet.get('name') or 'Sheet'
        safe = name[:31]
        for bad in ('\\', '/', '*', '?', '[', ']', ':'):
            safe = safe.replace(bad, '-')
        ws = wb.create_sheet(title=safe)

        group_headers = sheet.get('group_headers') or []
        headers = sheet.get('headers') or []
        rows = sheet.get('rows') or []

        # Row 1: group headers (if any), each group in a distinct color
        header_row_idx = 1
        if group_headers:
            for idx, (start_col, end_col, title) in enumerate(group_headers):
                color = group_palette[idx % len(group_palette)]
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                c = ws.cell(row=1, column=start_col, value=title)
                c.font = group_font
                c.fill = fill
                c.alignment = center
                c.border = box_border
                if end_col > start_col:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                # Fill merged siblings too (openpyxl merges don't auto-fill)
                for k in range(start_col + 1, end_col + 1):
                    c2 = ws.cell(row=1, column=k)
                    c2.fill = fill
                    c2.border = box_border
            # Paint uncovered base columns with a neutral dark bar so row 1 is visually complete
            covered = set()
            for (s, e, _t) in group_headers:
                for k in range(s, e + 1):
                    covered.add(k)
            neutral_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
            for j in range(1, len(headers) + 1):
                if j not in covered:
                    c = ws.cell(row=1, column=j, value='')
                    c.fill = neutral_fill
                    c.border = box_border
            ws.row_dimensions[1].height = 22
            header_row_idx = 2

        # Column headers row
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row_idx, column=j, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = box_border
        ws.row_dimensions[header_row_idx].height = 34

        # Data rows — date-shaped values become real Excel dates (dd-mm-yy).
        data_start = header_row_idx + 1
        for i, row in enumerate(rows):
            for j, v in enumerate(row, start=1):
                _write_cell(ws, data_start + i, j, v)

        # Auto-width
        for j in range(1, len(headers) + 1):
            try:
                col_letter = ws.cell(row=header_row_idx, column=j).column_letter
                max_len = len(str(ws.cell(row=header_row_idx, column=j).value or ''))
                for i in range(len(rows)):
                    v = ws.cell(row=data_start + i, column=j).value
                    max_len = max(max_len, len(str(v or '')))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
            except Exception:
                pass

        # Freeze top header rows
        ws.freeze_panes = ws.cell(row=data_start, column=1).coordinate

    if not wb.sheetnames:
        wb.create_sheet(title='Empty')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    if not filename.lower().endswith('.xlsx'):
        filename = filename.rsplit('.', 1)[0] + '.xlsx'
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )
